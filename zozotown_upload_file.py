# zozo_auto_upload.py

import pandas as pd
import os
import datetime as dt
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC

from selenium.webdriver.chrome.service import Service   # 自動Download
from selenium.webdriver.chrome.options import Options   # 自動Download
from webdriver_manager.chrome import ChromeDriverManager    # 自動Download

import platform
from openpyxl import load_workbook
from openpyxl.styles import Font

import re

from datetime import datetime


# ==============================
#  設定
# ==============================
# EXCEL_PATH = "\\160.251.168.21\Zozotown_List\List\ZOZOアップロード指示.xlsx"
EXCEL_PATH = ""
TARGET_SHEET = "作業カレンダー"
TEXT_DIR   = ""
# UPLOAD_DIR = "\\160.251.168.21\Zozotown_List\List_Data\"
UPLOAD_DIR = "\List_Data"

BASIC_USER = "zozotown-60"
BASIC_PASS = "z02o-tenant-0ff1ce"
FORM_USER  = "yasuda.k"
FORM_PASS  = ""
PASSWORD_FILE = ""

ZOZO_URL   = f"https://{BASIC_USER}:{BASIC_PASS}@to.zozo.jp/"

Up_list_name = []
Up_list_index = []
Up_list_cnt = 0
Error_flag = 0

STARTROW =  2     # EXcel 項目位置（日時	テキストファイ、、、、）

LOG_FILE = ""


# ==============================
# Main（拡張OK）
# ==============================
def zozotown_upload_file():
    """
    zozotown_upload_file()
    ├─ is_wsl()                  起動環境により設定値変更
    ├─ load_password()           パスワードをファイルから取得
    ├─ write_log()               ログ出力を行う
    ├─ read_excel()              Excel読み込み
    ├─ find_upload_file()        テキスト→CSV変換
    ├─ selenium_upload()         アップロード（失敗なら例外）
    ├─ update_excel_result()     Excel更新
    └─ update_excel_coller()     エラーの文字色を変更する
    """
    
    global EXCEL_PATH, TEXT_DIR   # ← 重要！　　更新する場合はGlobal必要
    global Up_list_cnt, Up_list_name, Up_list_index # 　更新する場合はGlobal必要
    global PASSWORD_FILE, FORM_PASS
    global LOG_FILE
    global Error_flag

    # 日付を YYYYMMDD形式で取得
    today_str = datetime.now().strftime("%Y%m%d")

    if is_wsl():  # テスト環境（WSL）
        EXCEL_PATH = "/mnt/z/List/ZOZOアップロード指示.xlsx"
        TEXT_DIR     = "/mnt/z/List_Data/"
        PASSWORD_FILE = "/mnt/z/Init/Password.txt"
        LOG_FILE = "/mnt/z/Log/" + f"{today_str}.txt"
    else:         # 本番環境（VPS Ubuntu）
        EXCEL_PATH   = "/srv/shared_zozo/List/ZOZOアップロード指示.xlsx"
        TEXT_DIR     = "/srv/shared_zozo/List_Data/"
        PASSWORD_FILE = "/srv/shared_zozo/Init/Password.txt"
        LOG_FILE = "/srv/shared_zozo/Log/" + f"{today_str}.txt"

    Up_list_name = []           # アップロードファイル名
    Up_list_index = []          # Excel内のIndex
    Up_list_cnt = 0             # UpLoadデータ件数
    Error_flag = 0              # EXcel 「エラー」色付けフラグ

    FORM_PASS = load_password(PASSWORD_FILE)    # パスワードを取得する

    df = read_excel()
    
    df = find_upload_file(df)

    if Up_list_cnt <= 0:
        print("🔸 アップロード対象ファイルがありません")
        write_log("🔸 アップロード対象ファイルがありません")
        return
    
    success = selenium_upload(df)

    if success:
        update_excel_result(df)

        print(f"Error_flag={Error_flag}")

        if Error_flag != 0:
            update_excel_coller()       # エラーのステータスを色付け
            write_log("アップロードエラー発生")


    print("☑ 全処理終了")


# ==============================
# ① Excel読み込み
# ==============================
def read_excel():
    
    global EXCEL_PATH, TEXT_DIR   # ← 重要！
    global STARTROW
    """
    Excel ファイルを読み込み、指定シートの DataFrame を返す。
    シートが存在しない場合は警告を出して None を返す。
    """

    try:
        # Excel のシート名一覧を取得
        all_sheets = pd.ExcelFile(EXCEL_PATH).sheet_names
        if TARGET_SHEET not in all_sheets:
            print(f"⚠ シート '{TARGET_SHEET}' が Excel ファイルに存在しません")
            write_log(f"⚠ シート '{TARGET_SHEET}' が Excel ファイルに存在しません")
            return None

        # 指定シートを読み込み　　A3からデータが保存されている
        df = pd.read_excel(EXCEL_PATH, sheet_name=TARGET_SHEET, header=STARTROW)
        # 念のためスペース除去
        # df.columns = df.columns.str.strip()
        # print(df.columns)  # ← 一度出力して確認してください

        return df

    except FileNotFoundError:
        print(f"❌ Excel ファイル '{EXCEL_PATH}' が存在しません")
        write_log(f"❌ Excel ファイル '{EXCEL_PATH}' が存在しません")
        return None

    except Exception as e:
        print(f"❌ Excel 読み込み時にエラー: {e}")
        write_log(f"❌ Excel 読み込み時にエラー: {e}")
        return None


# ==============================
# ② 対象ファイル抽出
# ==============================
def find_upload_file(df):

    global Up_list_name, Up_list_index, Up_list_cnt, Error_flag

    now = dt.datetime.now()

    for index, row in df.iterrows():
        # テスト
        # print("---- 行番号:", index, "----")  # 行番号
        # print(row.to_dict())                 # 行の内容（辞書形式）

        if pd.to_datetime(row["日時"]) <= now and str(row["処理結果"]) == "予約中":
            print(f">> {row.to_dict()}")  # ■テスト

            txt_name = f"{row['テキストファイル名']}.txt"
            txt_name = txt_name.strip()   # 改行や空白を除去
            txt_path = os.path.join(TEXT_DIR, txt_name)
            
            Up_list_name.append(txt_path)
            Up_list_index.append(index)

            if os.path.exists(txt_path):
                # 行数カウント（空白・改行のみの行は除外）
                with open(txt_path, "r", encoding="utf-8") as f:
                    line_count = sum(1 for line in f if line.strip())

                df.at[index, "データ行数表示"] = line_count
                df.at[index, "処理結果"] = "アップロード対象"
                Up_list_cnt += 1
            else:
                df.at[index, "処理結果"] = "エラー"
                df.at[index, "エラー情報"] = str(f"エラー ファイル無し: {txt_path}")

                Error_flag = -1
                write_log(f"エラー ファイル無し: {txt_path}")


            print(f'>> {df.at[index, "処理結果"]}')  # ■テスト

    return df

# ==============================
# ③ Seleniumアップロード
# ==============================
def selenium_upload(df):

    global Up_list_name, Up_list_index, Up_list_cnt, Error_flag
    global EXCEL_PATH, TEXT_DIR
    global PASSWORD_FILE, FORM_PASS

    print("Selenium開始:")

    options = Options()
    options.add_argument("--lang=ja-JP")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    if not is_wsl():  # 本番環境（VPS Ubuntu）
        options.add_argument("--headless")

    # ChromeDriver を自動で取得
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    # driver = webdriver.Chrome(options=options)

    try:
        driver.get(ZOZO_URL)
        wait = WebDriverWait(driver, 10)

        user_input = wait.until(EC.presence_of_element_located((By.ID, "UserID")))
        user_input.send_keys(FORM_USER)
        password_input = driver.find_element(By.NAME, "Password")
        password_input.send_keys(FORM_PASS + Keys.ENTER)
        
        for i in range(Up_list_cnt):
            file_path = Up_list_name[i]
            excel_idx = Up_list_index[i]

            print(f"アップロード　開始: {file_path}／{excel_idx}")     # テスト

            time.sleep(3)

            driver.get("https://to.zozo.jp/to/Advertisement.asp?c=RegistGoodsAd")
            file_input = wait.until(EC.presence_of_element_located((By.NAME, "upfile")))

            time.sleep(3)

            try:
                # ①　ファイルアップロード               
                file_input.send_keys(file_path)

                # ② ファイル名が Textエリアに表示されるまで待機
                wait = WebDriverWait(driver, 10)
                wait.until(
                    EC.text_to_be_present_in_element(
                        (By.CSS_SELECTOR, ".file-browser-value"),  # ファイル名表示エリア
                        os.path.basename(file_path)                # 表示されるファイル名
                    )
                )

                # time.sleep(100)

                # ③ 「アップロード」ボタンをクリック
                upload_button = driver.find_element(By.CSS_SELECTOR, ".zozoec-check")
                upload_button.click()

                # 結果が表示されるまで待機
                wait = WebDriverWait(driver, 10)
                result_element = wait.until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, ".form-primary"))
                )

                # テキスト取得
                result_text = result_element.text
                print("取得した結果:", result_text)

                # 正常登録件数を抽出
                match = re.search(r"正常登録件数：(\d+)件", result_text)
                if match:
                    count = int(match.group(1))
                    print("登録件数 =", count)
                    write_log(f"登録件数 = {count}")
                else:
                    print("件数を取得できませんでした")
                    write_log("件数を取得できませんでした")


                time.sleep(3)  # Upload完了待機
                print(f"アップロード　件数取得: {file_path}")

                if file_path.lower().endswith("on.txt"):
                    df.at[excel_idx, "ON登録数"] = count
                else:
                    df.at[excel_idx, "OFF登録数"] = count


                # Excel の処理結果を更新
                if count == 0:
                    df.at[excel_idx, "処理結果"] = "エラー"
                    df.at[excel_idx, "エラー情報"] = str("登録件数　０件　アップロードファイルを確認してください")
                    write_log("登録件数　０件　アップロードファイルを確認してください")
                    Error_flag = -1
                else:
                    df.at[excel_idx, "処理結果"] = "処理済み"
                    write_log("ステータスを「処理済み」にしました")

            except Exception as e:
                print(f"アップロード失敗: {file_path}, エラー: {e}")
                df.at[excel_idx, "処理結果"] = "エラー"
                df.at[excel_idx, "エラー情報"] = str(f"エラー アップロード失敗: {e}")
                Error_flag = -1
                write_log(f"アップロード失敗: {e}")


        time.sleep(3)
        print("アップロード完了:", file_path)
        write_log(f"アップロード完了: {file_path}")
        return True

    except Exception as e:
        print("Selenium エラー:", e)
        write_log(f"Selenium エラー: {e}")
        return False

    finally:
        driver.quit()


# ==============================
# ④ Excelへ結果反映
# ==============================
def update_excel_result(df, retries=3, wait_seconds=2):
    """
    df : DataFrame（A4から上書きしたい）
    retries : ファイルがロックされている場合のリトライ回数
    wait_seconds : リトライ時の待機秒数
    """
    global EXCEL_PATH, STARTROW
    
    attempt = 0
    while attempt < retries:
        try:
            # 既存ブックに追記・上書きする
            with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl",
                                mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name=TARGET_SHEET, index=False, startrow=STARTROW+1, header=False)
            
            print("✅ Excel更新完了")
            write_log("✅ Excel更新完了")
            return  # 成功したら抜ける

        except OSError as e:
            attempt += 1
            print(f"⚠️ Excelファイルがロック中またはアクセス不可: {e}")
            if attempt < retries:
                print(f"⏳ {wait_seconds}秒後にリトライします... ({attempt}/{retries})")
                time.sleep(wait_seconds)
            else:
                print("❌ Excelの更新に失敗しました。処理を中止します。")
                write_log("❌ Excelの更新に失敗しました。処理を中止します。")
                raise  # リトライしてもダメなら例外を上げる

    print("Excel更新完了")
    write_log("Excel更新完了")

# ==============================
# ⑤ ステータスの「エラー」文字を赤文字に変更
# ==============================
def update_excel_coller():
    global Up_list_name, Up_list_index, Up_list_cnt, Error_flag
    global EXCEL_PATH, TARGET_SHEET

    # Excel を開く
    wb = load_workbook(EXCEL_PATH)
    ws = wb[TARGET_SHEET]  # 対象シート

    red_font = Font(color="FF0000")   # 赤文字設定
    normal_font = Font(color="000000")  # 通常文字（黒）※必要な場合

    # --- アップしたファイルを1件ずつチェック ---
    for i in range(Up_list_cnt):
        row_idx = Up_list_index[i]    # ステータス位置（行番号）
        status_cell = ws.cell(row=row_idx+STARTROW+2, column=4)  # D列がステータス列

        # ステータスが「エラー」なら赤文字に変更
        if status_cell.value == "エラー":
            status_cell.font = red_font


    
    # 保存
    wb.save(EXCEL_PATH)
    wb.close()
    print("🎯 Excelステータスの色変更 完了")


# ==============================
#   グローバル設定 (環境別)
# ==============================
def is_wsl():
    # unameのreleaseに "microsoft" が含まれていれば WSL
    return 'microsoft' in platform.uname().release.lower()


# ==============================
#   パスワード取得 (環境別)
# ==============================
def load_password(file_path):
    """テキストファイルからパスワードを1行読み込む"""
    with open(file_path, "r", encoding="utf-8") as f:
        password = f.readline().strip()  # 改行を除く
    return password

# ==============================
#   ログ出力 (環境別)
# ==============================

def write_log(message):
    global LOG_FILE
    
    try:
        now = datetime.now().strftime("%H:%M:%S")
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"[{now}] -- {message}\n")

    except Exception as e:
        print("ログ書き込みエラー:", e)

if __name__ == "__main__":
    zozotown_upload_file()

